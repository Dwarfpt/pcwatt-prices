# -*- coding: utf-8 -*-
"""Builds an Excel report (stats.xlsx) from the scraped price feed:
components added and the stores they came from. Sheets:

  Summary      — totals at a glance
  By store     — items / categories / price ranges per shop
  By category  — items / store coverage / price ranges per part type
  By country   — stores and items per country

Usage:
  python build_stats_xlsx.py --feed ..\\..\\pcwatt-prices\\hardware_prices.csv --out stats.xlsx

EUR figures use the collector's rough EUR_RATES so cross-currency stores
are comparable; treat them as ballpark, not exact.
"""
import argparse
import csv
import datetime
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Reuse the collector's currency table and store->country map so the report
# stays consistent with what the app sees.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from price_collector import EUR_RATES, country_of_store  # noqa: E402

HEADER_FILL = PatternFill('solid', fgColor='1F2A37')
HEADER_FONT = Font(bold=True, color='FFFFFF')
TITLE_FONT = Font(bold=True, size=14)


def _eur(price, currency):
    return price / EUR_RATES.get(currency, 1.0)


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'A2'


def _autosize(ws):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 2, 60)


def load_rows(feed):
    with open(feed, encoding='utf-8-sig', newline='') as f:
        return list(csv.DictReader(f))


def build(feed, out):
    rows = load_rows(feed)
    wb = Workbook()

    # ── Summary ──
    ws = wb.active
    ws.title = 'Summary'
    stores = {}
    cats = {}
    countries = {}
    for r in rows:
        stores.setdefault(r['store'], []).append(r)
        cats.setdefault(r['category'], []).append(r)
        countries.setdefault(country_of_store(r['store']), []).append(r)
    ws['A1'] = 'PCWatt — статистика фида цен'
    ws['A1'].font = TITLE_FONT
    summary = [
        ('Сгенерировано', datetime.datetime.now().strftime('%Y-%m-%d %H:%M')),
        ('Всего позиций (компонентов)', len(rows)),
        ('Магазинов', len(stores)),
        ('Стран', len(countries)),
        ('Категорий', len(cats)),
        ('Валют', len({r['currency'] for r in rows})),
    ]
    for i, (k, v) in enumerate(summary, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 22

    # ── By store ──
    ws = wb.create_sheet('By store')
    ws.append(['Магазин', 'Страна', 'Валюта', 'Позиций', 'Категорий',
               'Средняя цена (EUR)', 'Дешевле всего (EUR)', 'Обновлён'])
    for store in sorted(stores, key=lambda s: -len(stores[s])):
        rs = stores[store]
        eurs = [_eur(float(r['price']), r['currency']) for r in rs]
        last = max((r['scraped_at'] for r in rs), default='')
        ws.append([
            store, country_of_store(store), rs[0]['currency'], len(rs),
            len({r['category'] for r in rs}),
            round(sum(eurs) / len(eurs)), round(min(eurs)), last[:10],
        ])
    _style_header(ws, 8)
    _autosize(ws)

    # ── By category ──
    ws = wb.create_sheet('By category')
    ws.append(['Категория', 'Позиций', 'Магазинов', 'Средняя (EUR)', 'Мин (EUR)', 'Макс (EUR)'])
    for cat in sorted(cats, key=lambda c: -len(cats[c])):
        rs = cats[cat]
        eurs = [_eur(float(r['price']), r['currency']) for r in rs]
        ws.append([cat, len(rs), len({r['store'] for r in rs}),
                   round(sum(eurs) / len(eurs)), round(min(eurs)), round(max(eurs))])
    _style_header(ws, 6)
    _autosize(ws)

    # ── By country ──
    ws = wb.create_sheet('By country')
    ws.append(['Страна', 'Магазинов', 'Позиций'])
    for country in sorted(countries, key=lambda c: -len(countries[c])):
        rs = countries[country]
        ws.append([country, len({r['store'] for r in rs}), len(rs)])
    _style_header(ws, 3)
    _autosize(ws)

    wb.save(out)
    print(f'[OK] {out} — {len(rows)} items, {len(stores)} stores, {len(countries)} countries')


if __name__ == '__main__':
    here = os.path.dirname(os.path.abspath(__file__))
    ap = argparse.ArgumentParser(description='Build stats.xlsx from the price feed')
    ap.add_argument('--feed', default=os.path.join(here, '..', '..', 'pcwatt-prices', 'hardware_prices.csv'))
    ap.add_argument('--out', default=os.path.join(here, '..', '..', 'pcwatt-prices', 'stats.xlsx'))
    args = ap.parse_args()
    build(args.feed, args.out)
